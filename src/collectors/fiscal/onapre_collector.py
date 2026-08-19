"""
Collector ONAPRE (Oficina Nacional de Presupuesto)
==================================================

Publica la ejecución presupuestaria y el presupuesto ciudadano como
XLS/PDF en su sitio. Estrategia:

1. ``fetch_report_links``: localiza enlaces a reportes (xls/xlsx/pdf) en la
   página de inicio.
2. ``parse_execution_xls``: parsea un XLS de ejecución (openpyxl) en
   ``BudgetExecution`` de forma tolerante (escanea encabezados y filas con
   montos).

Como el layout de las hojas cambia entre años, el parseo es deliberadamente
conservador: solo se emiten filas con monto numérico y etiqueta textual.
"""

import io
import logging
from typing import List, Optional

from bs4 import BeautifulSoup

from src.collectors.errors import CollectorSourceError
from src.collectors.http import http_get_bytes, http_get_text
from src.config import settings
from src.models.market import BudgetExecution

logger = logging.getLogger(__name__)

REPORT_EXTENSIONS = (".xls", ".xlsx", ".pdf")
REPORT_KEYWORDS = ("ejecucion", "ejecución", "presupuesto", "ciudadano")


def find_report_links(html: str, base_url: str) -> List[str]:
    """Enlaces absolutos a reportes presupuestarios en el HTML."""
    soup = BeautifulSoup(html, "html.parser")
    links = []
    for link in soup.find_all("a", href=True):
        href = link["href"]
        lower = href.lower()
        is_report = (
            lower.endswith(REPORT_EXTENSIONS)
            or any(k in lower for k in REPORT_KEYWORDS)
        )
        if not is_report:
            continue
        absolute = href if href.startswith("http") else base_url.rstrip("/") + "/" + href.lstrip("/")
        if absolute not in links:
            links.append(absolute)
    return links


def parse_execution_xls(content: bytes, year: int, url: Optional[str] = None) -> List[BudgetExecution]:
    """Convierte un XLS de ejecución presupuestaria en ``BudgetExecution``.

    Escanea todas las hojas; por fila toma el texto de la primera columna
    como entidad/partida y busca el primer monto numérico como ``amount_bs``.
    """
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - archivo corrupto o formato viejo
        raise CollectorSourceError(f"ONAPRE: XLS ilegible: {exc}") from exc

    items: List[BudgetExecution] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            values = [c.value for c in row]
            label = ""
            for v in values:
                if isinstance(v, str) and v.strip():
                    label = v.strip()
                    break
            if not label:
                continue
            amount = next(
                (float(v) for v in values if isinstance(v, (int, float)) and abs(v) > 0),
                None,
            )
            if amount is None:
                continue
            items.append(
                BudgetExecution(
                    year=year,
                    entity=label[:200],
                    amount_bs=amount,
                    url=url,
                )
            )
    return items


class ONAPRECollector:
    """Ejecución presupuestaria de la ONAPRE."""

    def __init__(self, base_url: Optional[str] = None):
        self.base_url = (base_url or settings.ONAPRE_BASE_URL).rstrip("/")

    def fetch_report_links(self) -> List[str]:
        """URLs de reportes presupuestarios en la página principal."""
        html = http_get_text(self.base_url + "/")
        return find_report_links(html, self.base_url)

    def collect(self, year: Optional[int] = None) -> List[BudgetExecution]:
        """Descarga el primer XLS de ejecución y lo parsea."""
        year = year or 2026
        links = self.fetch_report_links()
        xls_links = [url for url in links if url.lower().endswith((".xls", ".xlsx"))]
        if not xls_links:
            logger.warning("ONAPRE: sin archivos XLS en %s", self.base_url)
            return []

        content = http_get_bytes(xls_links[0])
        items = parse_execution_xls(content, year=year, url=xls_links[0])
        logger.info("ONAPRE %s: %d partidas parseadas", year, len(items))
        return items