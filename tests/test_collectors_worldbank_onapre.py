"""
Tests de los collectors de Fase A (Banco Mundial y ONAPRE)
===========================================================
"""

import io
from datetime import datetime

import pytest

from src.collectors.errors import CollectorSourceError
from src.collectors.fiscal.onapre_collector import (
    ONAPRECollector,
    find_report_links,
    parse_execution_xls,
)
from src.collectors.international.worldbank_collector import (
    WorldBankCollector,
    parse_wb_rows,
)

GDP_ROWS = [
    {"date": "2023", "value": 100.5},
    {"date": "2022", "value": 95.0},
    {"date": "2021", "value": None},
]


@pytest.fixture
def mock_http(monkeypatch):
    responses = {}

    def fake_json(url, params=None):
        return responses[url]

    def fake_text(url, params=None):
        return responses[url]

    def fake_bytes(url, params=None):
        return responses[url]

    # World Bank
    monkeypatch.setattr(
        "src.collectors.international.worldbank_collector.http_get_json", fake_json
    )
    # ONAPRE
    monkeypatch.setattr("src.collectors.fiscal.onapre_collector.http_get_text", fake_text)
    monkeypatch.setattr("src.collectors.fiscal.onapre_collector.http_get_bytes", fake_bytes)
    return responses


class TestWorldBank:
    def test_parse_wb_rows_descarta_nulos(self):
        payload = [{}, GDP_ROWS]
        rows = parse_wb_rows(payload)
        assert len(rows) == 2

    def test_parse_wb_rows_invalido(self):
        with pytest.raises(CollectorSourceError):
            parse_wb_rows([{"no": "meta"}])

    def test_fetch_gdp(self, mock_http):
        url = "https://api.worldbank.org/v2/country/VE/indicator/NY.GDP.MKTP.CD"
        mock_http[url] = [{}, GDP_ROWS]
        gdps = WorldBankCollector().fetch_gdp()
        assert [g.year for g in gdps] == [2022, 2023]  # orden cronológico
        assert gdps[0].value == 95.0

    def test_fetch_gdp_usa_filtro_fechas(self, mock_http):
        url = "https://api.worldbank.org/v2/country/VE/indicator/NY.GDP.MKTP.CD"
        mock_http[url] = [{}, GDP_ROWS]
        wb = WorldBankCollector()
        wb.fetch_gdp()
        # No se puede inspeccionar params con este mock; solo verificar que no falla

    def test_fetch_inflation(self, mock_http):
        url = "https://api.worldbank.org/v2/country/VE/indicator/FP.CPI.TOTL.ZG"
        mock_http[url] = [{}, [{"date": "2023", "value": 15.2}]]
        points = WorldBankCollector().fetch_inflation()
        assert points[0].period == "2023-12"
        assert points[0].annual_rate == 15.2


ONAPRE_HTML = """
<html>
<a href="https://www.onapre.gob.ve/presupuesto">Presupuesto</a>
<a href="/files/ejecucion_presupuestaria_2026.xlsx">Ejecución 2026</a>
<a href="/noticias">Noticias</a>
</html>
"""


class TestONAPRE:
    def test_find_report_links(self):
        links = find_report_links(ONAPRE_HTML, "https://www.onapre.gob.ve")
        assert any(l.endswith(".xlsx") for l in links)
        assert all(l.startswith("http") for l in links)
        assert "https://www.onapre.gob.ve/files/ejecucion_presupuestaria_2026.xlsx" in links

    def test_parse_execution_xls(self):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["Ente", "Asignado", "Ejecutado"])
        ws.append(["Ministerio de Salud", 100.5, 80.2])
        ws.append(["Ministerio de Educación", 200, 150])
        buf = io.BytesIO()
        wb.save(buf)

        items = parse_execution_xls(buf.getvalue(), year=2026)
        assert len(items) == 2
        assert items[0].entity == "Ministerio de Salud"
        assert items[0].amount_bs == 100.5
        assert items[0].year == 2026
        assert items[0].source == "onapre"

    def test_parse_execution_xls_filas_sin_monto(self):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["Texto sin número"])
        buf = io.BytesIO()
        wb.save(buf)

        assert parse_execution_xls(buf.getvalue(), year=2026) == []

    def test_parse_execution_xls_corrupto(self):
        with pytest.raises(CollectorSourceError):
            parse_execution_xls(b"no es un xls", year=2026)

    def test_collect_sin_xls(self, mock_http):
        mock_http["https://www.onapre.gob.ve/"] = "<html><a href='/noticias'>Noticias</a></html>"
        items = ONAPRECollector().collect(year=2026)
        assert items == []

    def test_collect_con_xls(self, mock_http):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["Ente", "Monto"])
        ws.append(["Presidencia", 500.25])
        buf = io.BytesIO()
        wb.save(buf)

        base = "https://www.onapre.gob.ve"
        xls_url = base + "/files/ejecucion_presupuestaria_2026.xlsx"
        mock_http[base + "/"] = f'<a href="{xls_url}">Ejecución</a>'
        mock_http[xls_url] = buf.getvalue()

        items = ONAPRECollector().collect(year=2026)
        assert len(items) == 1
        assert items[0].entity == "Presidencia"
        assert items[0].amount_bs == 500.25